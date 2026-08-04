"""Runtime dataset preflight: verifies source paths, expected sheets/columns and
output permissions before the graph runs, instead of hardcoding dataset assumptions.
"""
from __future__ import annotations

import glob
import os
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from src.config.runtime_config import RuntimeCafeConfig

# Expected columns per source, taken from the exact supplied schemas (plan section 9).
_EXPECTED_COLUMNS = {
    "pos": {
        "transaction_id", "timestamp", "sku", "item_name", "quantity",
        "unit_price_sar", "discount_sar", "line_total_sar", "payment_method",
        "channel", "cashier_id",
    },
    "menu": {
        "sku", "item_en", "item_ar", "category", "price_sar", "unit_cost_sar",
        "is_iced", "launch_date", "retire_date",
    },
    "traffic": {"date", "hour", "door_count"},
    "staff": {
        "date", "employee_id", "name", "role", "shift_start", "shift_end",
        "hours", "hourly_rate_sar",
    },
}

_REQUIRED_ENV_VARS = ["OPENAI_API_KEY"]
_OPTIONAL_ENV_VARS = ["TAVILY_API_KEY", "LANGCHAIN_API_KEY"]


@dataclass
class PreflightReport:
    ok: bool
    source_checks: dict[str, dict] = field(default_factory=dict)
    missing_env: list[str] = field(default_factory=list)
    missing_optional_env: list[str] = field(default_factory=list)
    output_writable: bool = True
    errors: list[str] = field(default_factory=list)


def _check_csv(path: Path, expected_cols: set[str]) -> dict:
    if not path.exists():
        return {"status": "missing", "path": str(path)}
    try:
        df = pd.read_csv(path, nrows=5)
    except Exception as e:  # noqa: BLE001
        return {"status": "unreadable", "path": str(path), "error": str(e)}
    missing = expected_cols - set(df.columns)
    return {
        "status": "ok" if not missing else "schema_mismatch",
        "path": str(path),
        "missing_columns": sorted(missing),
    }


def _check_inventory(path: Path, sheet: str) -> dict:
    if not path.exists():
        return {"status": "missing", "path": str(path)}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"status": "unreadable", "path": str(path), "error": str(e)}
    if sheet not in wb.sheetnames:
        return {"status": "missing_sheet", "path": str(path), "sheets_found": wb.sheetnames}
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = {"week_starting", "sku", "item", "units_ordered", "units_sold", "units_wasted", "unit_cost_sar"}
    missing = expected - set(header)
    return {
        "status": "ok" if not missing else "schema_mismatch",
        "path": str(path),
        "missing_columns": sorted(missing),
    }


def _check_reviews(path: Path) -> dict:
    if not path.exists():
        return {"status": "missing", "path": str(path)}
    import json

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        return {"status": "unreadable", "path": str(path), "error": str(e)}
    if not isinstance(data, list):
        return {"status": "schema_mismatch", "path": str(path)}
    expected = {"review_id", "date", "source", "rating", "text"}
    missing = expected - set(data[0].keys()) if data else set()
    return {
        "status": "ok" if not missing else "schema_mismatch",
        "path": str(path),
        "record_count": len(data),
        "missing_fields": sorted(missing),
    }


def _check_emails(pattern: str, data_dir: Path) -> dict:
    files = glob.glob(str(data_dir / pattern))
    return {"status": "ok" if files else "missing", "count": len(files), "pattern": pattern}


def run_preflight(config: RuntimeCafeConfig) -> PreflightReport:
    report = PreflightReport(ok=True)

    for source in config.source_registry.sources:
        path = config.data_dir / source.path
        if source.name in ("pos", "menu", "traffic", "staff"):
            check = _check_csv(path, _EXPECTED_COLUMNS[source.name])
        elif source.name == "inventory":
            check = _check_inventory(path, source.sheet or "weekly_counts")
        elif source.name == "reviews":
            check = _check_reviews(path)
        elif source.name == "emails":
            check = _check_emails(source.path, config.data_dir)
        else:
            check = {"status": "unknown_source"}
        report.source_checks[source.name] = check
        if check["status"] not in ("ok",):
            report.errors.append(f"{source.name}: {check['status']}")

    for var in _REQUIRED_ENV_VARS:
        if not os.environ.get(var):
            report.missing_env.append(var)
    for var in _OPTIONAL_ENV_VARS:
        if not os.environ.get(var):
            report.missing_optional_env.append(var)

    try:
        config.artifact_root.mkdir(parents=True, exist_ok=True)
        config.checkpoint_db.parent.mkdir(parents=True, exist_ok=True)
        config.memory_db.parent.mkdir(parents=True, exist_ok=True)
        probe = config.artifact_root / ".preflight_write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink()
    except OSError as e:
        report.output_writable = False
        report.errors.append(f"output directory not writable: {e}")

    # A missing critical source (any of the 7) is not fatal here; the ingestion
    # graph handles per-source failure independently. Preflight fails hard only
    # when *all* sources are unreadable or output cannot be written, or a
    # required secret is absent.
    all_missing = all(c["status"] == "missing" for c in report.source_checks.values())
    report.ok = (
        not all_missing
        and report.output_writable
        and not report.missing_env
    )
    return report
